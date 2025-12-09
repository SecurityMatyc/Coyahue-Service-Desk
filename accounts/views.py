from datetime import timedelta
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden, HttpResponse
from django.contrib import messages
from django.urls import reverse
from datetime import timedelta
from django.db.models import Count, Avg, Q
from django.db.models.functions import TruncDate, TruncMonth
from django.utils import timezone

from django.db.models import Count, Avg, DurationField, ExpressionWrapper, F, Q
from django.db.models.functions import TruncDay
from django.utils import timezone
from django.template.loader import get_template

from rest_framework import generics
from rest_framework.permissions import AllowAny

from xhtml2pdf import pisa

from .models import Usuario, Rol, Tecnico
from .serializers import RegistroUsuarioSerializer, UsuarioSerializer

from tickets.models import (
    Ticket,
    Categoria,
    Subcategoria,
    Prioridad,
    EstadoTicket,
    AreaAfectada,
    AsignacionTicket,
    HistorialTicket,
    ComentarioTicket,
    CalificacionTicket,
)

from notifications.models import Notificacion


# -------------------------------------------------------------------
# Helper de rol
# -------------------------------------------------------------------

def require_role(user, rol_nombre: str) -> bool:
    return (
        user.is_authenticated
        and hasattr(user, "rol")
        and user.rol
        and user.rol.nombre_rol.upper() == rol_nombre.upper()
    )


# -------------------------------------------------------------------
# RECUPERAR CONTRASEÑA
# -------------------------------------------------------------------

def recuperar_contrasena(request):
    """
    Permite a usuarios y técnicos solicitar recuperación de contraseña.
    Crea notificación para administradores.
    """
    if request.method == "POST":
        email = request.POST.get("email", "").strip().lower()

        if not email:
            messages.error(request, "Por favor ingresa tu email.")
            return redirect("recuperar_contrasena")

        # Buscar usuario por email
        try:
            usuario = Usuario.objects.get(email=email)
            
            # Verificar que no sea admin
            if usuario.rol and usuario.rol.nombre_rol.upper() == "ADMIN":
                messages.info(request, "Los administradores deben contactar al equipo de TI directamente.")
                return redirect("login")

            # Crear notificación para todos los admins
            admins = Usuario.objects.filter(rol__nombre_rol__iexact="ADMIN")
            
            for admin in admins:
                Notificacion.objects.create(
                    ticket=None,  # No hay ticket asociado
                    usuario_destino=admin,
                    tipo_notificacion="recuperacion_password",
                    titulo=f"🔐 Solicitud de recuperación de contraseña",
                    mensaje=f"{usuario.get_full_name() or usuario.email} solicita resetear su contraseña.",
                    canal_notificacion="portal"
                )

            messages.success(
                request,
                "✅ Solicitud enviada exitosamente! Si el email está registrado, el administrador te contactará pronto."
            )
            return render(request, "recuperar_contrasena.html")

        except Usuario.DoesNotExist:
            # Mensaje genérico para no revelar si el email existe o no (seguridad)
            messages.success(
                request,
                "✅ Solicitud enviada exitosamente! Si el email está registrado, el administrador te contactará pronto."
            )
            return render(request, "recuperar_contrasena.html")

    return render(request, "recuperar_contrasena.html")


# -------------------------------------------------------------------
# API (DRF)
# -------------------------------------------------------------------

class RegistroUsuarioView(generics.CreateAPIView):
    queryset = Usuario.objects.all()
    serializer_class = RegistroUsuarioSerializer
    permission_classes = [AllowAny]


class PerfilUsuarioView(generics.RetrieveAPIView):
    serializer_class = UsuarioSerializer

    def get_object(self):
        return self.request.user


# -------------------------------------------------------------------
# LOGIN
# -------------------------------------------------------------------

def login_view(request):
    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password")

        user = authenticate(request, username=email, password=password)

        if user is not None:
            login(request, user)
            rol = getattr(user.rol, "nombre_rol", "").upper()

            if rol == "ADMIN":
                return redirect("dashboard_admin")
            elif rol == "TECNICO":
                return redirect("dashboard_tecnico")
            else:
                return redirect("dashboard_usuario")

        return render(request, "login.html", {"form": {"errors": True}})

    return render(request, "login.html")


# -------------------------------------------------------------------
# DASHBOARD ADMIN
# -------------------------------------------------------------------

@login_required
def dashboard_admin(request):
    if not require_role(request.user, "ADMIN"):
        return HttpResponseForbidden("No tienes permiso para ver este panel.")

    total_usuarios = Usuario.objects.count()
    total_tickets = Ticket.objects.count()

    tickets_abiertos = Ticket.objects.filter(estado__nombre_estado="Abierto").count()
    tickets_en_progreso = Ticket.objects.filter(estado__nombre_estado="En Progreso").count()
    tickets_resueltos = Ticket.objects.filter(estado__nombre_estado="Resuelto").count()
    tickets_cerrados = Ticket.objects.filter(estado__nombre_estado="Cerrado").count()

    tickets_por_prioridad = (
        Prioridad.objects
        .annotate(total=Count("tickets"))
        .order_by("nivel")
    )

    tickets_por_area = (
        AreaAfectada.objects
        .annotate(total=Count("tickets"))
        .order_by("nombre_area")
    )

    ultimos_tickets = (
        Ticket.objects
        .select_related("solicitante", "estado", "prioridad", "area_afectada")
        .order_by("-fecha_creacion")[:5]
    )

    # Notificaciones para la campanita
    notificaciones_no_leidas = (
        Notificacion.objects
        .filter(usuario_destino=request.user, leida=False)
        .order_by("-fecha_envio")[:5]
    )
    total_notif_no_leidas = Notificacion.objects.filter(
        usuario_destino=request.user,
        leida=False,
    ).count()

    context = {
        "total_usuarios": total_usuarios,
        "total_tickets": total_tickets,
        "tickets_abiertos": tickets_abiertos,
        "tickets_en_progreso": tickets_en_progreso,
        "tickets_resueltos": tickets_resueltos,
        "tickets_cerrados": tickets_cerrados,
        "tickets_por_prioridad": tickets_por_prioridad,
        "tickets_por_area": tickets_por_area,
        "ultimos_tickets": ultimos_tickets,
        "ahora": timezone.now(),
        "notificaciones_no_leidas": notificaciones_no_leidas,
        "total_notif_no_leidas": total_notif_no_leidas,
    }

    return render(request, "admin/dashboard.html", context)


# -------------------------------------------------------------------
# DASHBOARD TÉCNICO
# -------------------------------------------------------------------

@login_required
def dashboard_tecnico(request):
    if not require_role(request.user, "TECNICO"):
        return HttpResponseForbidden("No tienes permiso para ver este panel.")

    tecnico = getattr(request.user, "tecnico", None)
    if tecnico is None:
        tecnico, _ = Tecnico.objects.get_or_create(usuario=request.user)

    tickets_asignados = Ticket.objects.filter(
        asignaciones__tecnico_asignado=tecnico,
        asignaciones__activo=True,
    ).select_related("categoria", "estado", "prioridad").distinct()

    tickets_abiertos = tickets_asignados.filter(estado__nombre_estado="Abierto").count()
    tickets_en_progreso = tickets_asignados.filter(estado__nombre_estado="En progreso").count()
    tickets_resueltos = tickets_asignados.filter(estado__nombre_estado="Resuelto").count()

    return render(request, "tecnico/dashboard_tecnico.html", {
        "tecnico": tecnico,
        "tickets_asignados": tickets_asignados,
        "tickets_abiertos": tickets_abiertos,
        "tickets_en_progreso": tickets_en_progreso,
        "tickets_resueltos": tickets_resueltos,
    })


# -------------------------------------------------------------------
# DASHBOARD USUARIO SOLICITANTE
# -------------------------------------------------------------------

@login_required
def dashboard_usuario(request):
    if not require_role(request.user, "USUARIO"):
        return HttpResponseForbidden("No tienes permiso para ver este panel.")

    tickets = Ticket.objects.filter(
        solicitante=request.user
    ).select_related("estado", "categoria", "prioridad")

    total = tickets.count()
    abiertos = tickets.filter(estado__nombre_estado="Abierto").count()
    en_progreso = tickets.filter(estado__nombre_estado="En progreso").count()
    resueltos = tickets.filter(estado__nombre_estado="Resuelto").count()

    ultimos = tickets.order_by("-fecha_creacion")[:5]

    return render(request, "usuario/dashboard_usuario.html", {
        "total": total,
        "abiertos": abiertos,
        "en_progreso": en_progreso,
        "resueltos": resueltos,
        "ultimos": ultimos,
    })


# -------------------------------------------------------------------
# ADMIN: CRUD Usuarios
# -------------------------------------------------------------------

@login_required
def usuarios_listar(request):
    if not require_role(request.user, "ADMIN"):
        return HttpResponseForbidden("No tienes permiso para ver esta sección.")

    usuarios = Usuario.objects.select_related("rol").all()
    return render(request, "admin/usuarios_listar.html", {"usuarios": usuarios})


@login_required
def usuarios_crear(request):
    if not require_role(request.user, "ADMIN"):
        return HttpResponseForbidden("No tienes permiso para ver esta sección.")

    roles = Rol.objects.all()

    if request.method == "POST":
        email = request.POST["email"]
        first_name = request.POST.get("first_name")
        last_name = request.POST.get("last_name")
        password = request.POST["password"]
        rol_id = request.POST["rol"]

        Usuario.objects.create_user(
            email=email,
            first_name=first_name,
            last_name=last_name,
            password=password,
            rol_id=rol_id,
        )
        return redirect("usuarios_listar")

    return render(request, "admin/usuarios_crear.html", {"roles": roles})


@login_required
def usuarios_editar(request, usuario_id):
    if not require_role(request.user, "ADMIN"):
        return HttpResponseForbidden("No tienes permiso para ver esta sección.")

    usuario = get_object_or_404(Usuario, id=usuario_id)
    roles = Rol.objects.all()

    if request.method == "POST":
        usuario.email = request.POST["email"]
        usuario.first_name = request.POST.get("first_name")
        usuario.last_name = request.POST.get("last_name")
        usuario.rol_id = request.POST["rol"]
        
        # Cambiar contraseña solo si se proporcionó una nueva
        nueva_password = request.POST.get("nueva_password", "").strip()
        if nueva_password:
            if len(nueva_password) < 8:
                messages.error(request, "La contraseña debe tener al menos 8 caracteres.")
                return render(request, "admin/usuarios_editar.html", {
                    "usuario": usuario,
                    "roles": roles,
                })
            usuario.set_password(nueva_password)
            messages.success(request, f"Usuario actualizado. Contraseña cambiada correctamente.")
        else:
            messages.success(request, "Usuario actualizado correctamente.")
        
        usuario.save()
        return redirect("usuarios_listar")

    return render(request, "admin/usuarios_editar.html", {
        "usuario": usuario,
        "roles": roles,
    })


@login_required
def usuarios_eliminar(request, usuario_id):
    if not require_role(request.user, "ADMIN"):
        return HttpResponseForbidden("No tienes permiso para ver esta sección.")

    usuario = get_object_or_404(Usuario, id=usuario_id)

    if request.method == "POST":
        usuario.delete()
        return redirect("usuarios_listar")

    return render(request, "admin/usuarios_eliminar.html", {
        "usuario": usuario
    })


# -------------------------------------------------------------------
# ADMIN: Listados de catálogos (solo lectura)
# -------------------------------------------------------------------

@login_required
def roles_listar(request):
    if not require_role(request.user, "ADMIN"):
        return HttpResponseForbidden("No tienes permiso para ver esta sección.")

    roles = Rol.objects.all()
    return render(request, "admin/roles_listar.html", {"roles": roles})


@login_required
def categorias_listar(request):
    if not require_role(request.user, "ADMIN"):
        return HttpResponseForbidden("No tienes permiso para ver esta sección.")

    categorias = Categoria.objects.all()
    return render(request, "admin/categorias_listar.html", {"categorias": categorias})


@login_required
def subcategorias_listar(request):
    if not require_role(request.user, "ADMIN"):
        return HttpResponseForbidden("No tienes permiso para ver esta sección.")

    subcategorias = Subcategoria.objects.select_related("categoria")
    return render(request, "admin/subcategorias_listar.html", {
        "subcategorias": subcategorias
    })


@login_required
def prioridades_listar(request):
    if not require_role(request.user, "ADMIN"):
        return HttpResponseForbidden("No tienes permiso para ver esta sección.")

    prioridades = Prioridad.objects.all()
    return render(request, "admin/prioridades_listar.html", {
        "prioridades": prioridades
    })


@login_required
def estados_listar(request):
    if not require_role(request.user, "ADMIN"):
        return HttpResponseForbidden("No tienes permiso para ver esta sección.")

    estados = EstadoTicket.objects.all()
    return render(request, "admin/estados_listar.html", {
        "estados": estados
    })


# -------------------------------------------------------------------
# Export / Reportes
# -------------------------------------------------------------------
def render_to_pdf(template_src, context_dict=None):
    """
    Renderiza un template HTML a PDF usando xhtml2pdf.
    """
    if context_dict is None:
        context_dict = {}

    template = get_template(template_src)
    html = template.render(context_dict)

    result = io.BytesIO()
    pdf = pisa.pisaDocument(
        io.BytesIO(html.encode("UTF-8")),
        dest=result,
        encoding="UTF-8",
    )

    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type="application/pdf")
    return None

# -----------------------------------------------
# REPORTES ADMIN (Dashboard Ejecutivo / Táctico / Estratégico)
# -----------------------------------------------
from datetime import timedelta
from django.db.models import Count, Avg, Q
from django.db.models.functions import TruncDate, TruncMonth
from django.utils import timezone

@login_required
def reportes_dashboard(request):
    if not require_role(request.user, "ADMIN"):
        return HttpResponseForbidden("No tienes permiso para ver esta sección.")

    hoy = timezone.now().date()
    ahora = timezone.now()
    hace_7 = hoy - timedelta(days=7)
    hace_30 = hoy - timedelta(days=30)
    hace_180 = hoy - timedelta(days=180)

    # ----------------------------
    # Estados finales / no finales
    # ----------------------------
    estados_finales_ids = list(
        EstadoTicket.objects.filter(es_final=True).values_list("id", flat=True)
    )
    estados_no_finales_ids = list(
        EstadoTicket.objects.filter(es_final=False).values_list("id", flat=True)
    )

    tickets = Ticket.objects.select_related("estado", "prioridad", "area_afectada")

    # ----------------------------
    # Backlog y comparación con ayer
    # ----------------------------
    backlog_qs = tickets.exclude(estado_id__in=estados_finales_ids)
    backlog_total = backlog_qs.count()

    backlog_ayer = (
        Ticket.objects.filter(fecha_creacion__date__lte=hoy - timedelta(days=1))
        .exclude(estado_id__in=estados_finales_ids)
        .count()
    )
    backlog_vs_ayer = backlog_total - backlog_ayer

    # Edad del backlog (en horas y máximo en días)
    backlog_con_fecha = backlog_qs.filter(fecha_creacion__isnull=False)
    backlog_edad_promedio_horas = None
    backlog_edad_max_dias = None
    if backlog_con_fecha.exists():
        edades_horas = [
            (ahora - t.fecha_creacion).total_seconds() / 3600.0
            for t in backlog_con_fecha
        ]
        backlog_edad_promedio_horas = round(sum(edades_horas) / len(edades_horas), 1)
        backlog_edad_max_dias = round(max(edades_horas) / 24.0, 1)

    # ----------------------------
    # Throughput últimos 7 días (tickets cerrados)
    # ----------------------------
    tickets_resueltos_7 = Ticket.objects.filter(
        estado_id__in=estados_finales_ids,
        fecha_cierre__date__gte=hace_7,
    )
    throughput_7 = tickets_resueltos_7.count()

    # ----------------------------
    # SLA cumplimiento
    # ----------------------------
    tickets_con_sla = Ticket.objects.filter(
        sla_horas_objetivo__isnull=False, fecha_cierre__isnull=False
    )
    dentro_sla = 0
    for t in tickets_con_sla:
        if t.fecha_creacion and t.fecha_cierre:
            diff_horas = (t.fecha_cierre - t.fecha_creacion).total_seconds() / 3600.0
            if diff_horas <= t.sla_horas_objetivo:
                dentro_sla += 1

    sla_porcentaje = (
        round(dentro_sla / tickets_con_sla.count() * 100, 1)
        if tickets_con_sla.exists()
        else None
    )

    # SLA en rojo si < 80%
    sla_rojo = sla_porcentaje is not None and sla_porcentaje < 80

    # ----------------------------
    # MTTR (Mean Time To Resolve) últimos 30 días
    # ----------------------------
    cerrados_30 = Ticket.objects.filter(
        fecha_cierre__date__gte=hace_30,
        fecha_creacion__isnull=False,
        fecha_cierre__isnull=False,
    )
    mttr_horas_30 = None
    if cerrados_30.exists():
        diffs_horas = [
            (t.fecha_cierre - t.fecha_creacion).total_seconds() / 3600.0
            for t in cerrados_30
        ]
        mttr_horas_30 = round(sum(diffs_horas) / len(diffs_horas), 1)

    # ----------------------------
    # Tickets críticos pendientes
    # ----------------------------
    criticos_pendientes = Ticket.objects.filter(
        prioridad__nombre_prioridad__iexact="Crítica"
    ).exclude(estado_id__in=estados_finales_ids).count()
    criticos_rojo = criticos_pendientes > 0

    # ----------------------------
    # Reaperturas (proxy de calidad)
    # ----------------------------
    reaperturas = HistorialTicket.objects.filter(
        estado_anterior__id__in=estados_finales_ids,
        estado_nuevo__nombre_estado__iexact="Abierto",
    ).count()

    # ----------------------------
    # CFD (Cumulative Flow) últimos 7 días
    # ----------------------------
    cfd_qs = (
        Ticket.objects.filter(fecha_creacion__date__gte=hace_7)
        .annotate(dia=TruncDate("fecha_creacion"))
        .values("dia", "estado__nombre_estado")
        .annotate(total=Count("id"))
        .order_by("dia")
    )
    cfd = list(cfd_qs)

    # ----------------------------
    # Burndown: backlog por día (últimos 7 días)
    # ----------------------------
    burndown = []
    for i in range(6, -1, -1):
        dia = hoy - timedelta(days=i)
        pendientes_en_dia = (
            Ticket.objects.filter(fecha_creacion__date__lte=dia)
            .exclude(
                Q(estado_id__in=estados_finales_ids) & Q(fecha_cierre__date__lte=dia)
            )
            .count()
        )
        burndown.append({"dia": dia, "backlog": pendientes_en_dia})

    # ----------------------------
    # Distribución de carga del equipo (por Técnico)
    # ----------------------------
    carga_tecnicos = (
        Tecnico.objects.annotate(
            activos=Count(
                "tickets_asignados",
                filter=Q(
                    tickets_asignados__ticket__estado_id__in=estados_no_finales_ids
                ),
                distinct=True,
            ),
            resueltos=Count(
                "tickets_asignados",
                filter=Q(
                    tickets_asignados__ticket__estado_id__in=estados_finales_ids
                ),
                distinct=True,
            ),
        )
        .select_related("usuario")
        .order_by("-activos")
    )

    # ----------------------------
    # Top 5 categorías (últimos 30 días)
    # ----------------------------
    top_categorias = (
        Categoria.objects.annotate(
            total=Count(
                "tickets",
                filter=Q(tickets__fecha_creacion__date__gte=hace_30),
            )
        )
        .order_by("-total")[:5]
    )

    # ----------------------------
    # Tendencias mensuales (últimos 6 meses)
    # ----------------------------
    tendencias_qs = (
        Ticket.objects.filter(fecha_creacion__date__gte=hace_180)
        .annotate(mes=TruncMonth("fecha_creacion"))
        .values("mes")
        .annotate(total=Count("id"))
        .order_by("mes")
    )
    tendencias = list(tendencias_qs)

    # Pronóstico simple = promedio de los últimos 3 meses
    if len(tendencias) >= 1:
        ultimos_3 = tendencias[-3:]
        forecast = round(sum(m["total"] for m in ultimos_3) / len(ultimos_3))
    else:
        forecast = None

    # ----------------------------
    # Alertas visuales
    # ----------------------------
    backlog_rojo = backlog_total > 30  # umbral ajustable
    backlog_edad_rojo = (
        backlog_edad_promedio_horas is not None and backlog_edad_promedio_horas > 48
    )

    # ----------------------------
    # CSAT (Satisfacción del cliente)
    # ----------------------------
    calificaciones = CalificacionTicket.objects.all()
    csat_promedio = None
    csat_porcentaje = None
    csat_resueltos_porcentaje = None
    csat_total_calificaciones = calificaciones.count()

    if calificaciones.exists():
        # Promedio de puntuación (1-5 estrellas)
        suma_puntuaciones = sum(c.puntuacion for c in calificaciones)
        csat_promedio = round(suma_puntuaciones / csat_total_calificaciones, 1)
        csat_porcentaje = round((csat_promedio / 5) * 100, 1)  # Convertir a %
        
        # % de tickets donde el problema fue resuelto
        resueltos_count = calificaciones.filter(resuelto=True).count()
        csat_resueltos_porcentaje = round((resueltos_count / csat_total_calificaciones) * 100, 1)

    return render(request, "reportes/dashboard.html", {
        # Vista ejecutiva
        "backlog_total": backlog_total,
        "backlog_vs_ayer": backlog_vs_ayer,
        "backlog_edad_promedio_horas": backlog_edad_promedio_horas,
        "backlog_edad_max_dias": backlog_edad_max_dias,
        "throughput_7": throughput_7,
        "sla_porcentaje": sla_porcentaje,
        "mttr_horas_30": mttr_horas_30,
        "criticos_pendientes": criticos_pendientes,
        "reaperturas": reaperturas,

        # CSAT
        "csat_promedio": csat_promedio,
        "csat_porcentaje": csat_porcentaje,
        "csat_resueltos_porcentaje": csat_resueltos_porcentaje,
        "csat_total_calificaciones": csat_total_calificaciones,

        # Alertas
        "sla_rojo": sla_rojo,
        "criticos_rojo": criticos_rojo,
        "backlog_rojo": backlog_rojo,
        "backlog_edad_rojo": backlog_edad_rojo,

        # Táctico
        "carga_tecnicos": carga_tecnicos,
        "top_categorias": top_categorias,

        # Estratégico / gráficos
        "cfd": cfd,
        "burndown": burndown,
        "tendencias": tendencias,
        "forecast": forecast,
    })

@login_required
def reportes_tickets_excel(request):
    if not require_role(request.user, "ADMIN"):
        return HttpResponseForbidden("No tienes permiso para exportar.")

    # Crear workbook y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"

    # --- ENCABEZADO PRINCIPAL ---
    ws.merge_cells('A1:I1')
    cell_titulo = ws['A1']
    cell_titulo.value = "COYAHUE SERVICE DESK - REPORTE DE TICKETS"
    cell_titulo.font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
    cell_titulo.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cell_titulo.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Info generación
    fecha_generacion = timezone.localtime(timezone.now())
    ws['A2'] = f"Generado por: {request.user.email}"
    ws['A3'] = f"Fecha: {fecha_generacion.strftime('%d-%m-%Y %H:%M')}"
    ws['A2'].font = Font(size=9, italic=True)
    ws['A3'].font = Font(size=9, italic=True)

    # --- ENCABEZADOS DE TABLA ---
    headers = [
        "ID", "Título", "Solicitante", "Estado", "Prioridad",
        "Área", "Fecha creación", "Fecha cierre", "Creado"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # --- DATOS ---
    tickets = Ticket.objects.select_related(
        "solicitante", "estado", "prioridad", "area_afectada"
    ).order_by("-fecha_creacion")

    row_num = 6
    for ticket in tickets:
        # Datos del ticket
        ws.cell(row=row_num, column=1, value=ticket.id)
        ws.cell(row=row_num, column=2, value=ticket.titulo)
        ws.cell(row=row_num, column=3, value=ticket.solicitante.email if ticket.solicitante else "")
        ws.cell(row=row_num, column=4, value=ticket.estado.nombre_estado if ticket.estado else "")
        ws.cell(row=row_num, column=5, value=ticket.prioridad.nombre_prioridad if ticket.prioridad else "-")
        ws.cell(row=row_num, column=6, value=ticket.area_afectada.nombre_area if ticket.area_afectada else "-")
        
        # Convertir fechas a zona horaria local antes de mostrar
        fecha_creacion_local = timezone.localtime(ticket.fecha_creacion) if ticket.fecha_creacion else None
        fecha_cierre_local = timezone.localtime(ticket.fecha_cierre) if ticket.fecha_cierre else None
        
        ws.cell(row=row_num, column=7, value=fecha_creacion_local.strftime("%d-%m-%Y %H:%M") if fecha_creacion_local else "")
        ws.cell(row=row_num, column=8, value=fecha_cierre_local.strftime("%d-%m-%Y %H:%M") if fecha_cierre_local else "No cerrado")
        ws.cell(row=row_num, column=9, value=fecha_creacion_local.strftime("%d-%m-%Y") if fecha_creacion_local else "")

        # Aplicar colores según estado
        estado_cell = ws.cell(row=row_num, column=4)
        if ticket.estado:
            estado_nombre = ticket.estado.nombre_estado.lower()
            if "cerrado" in estado_nombre or "resuelto" in estado_nombre:
                estado_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde claro
                estado_cell.font = Font(color="006100")
            elif "progreso" in estado_nombre:
                estado_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Amarillo claro
                estado_cell.font = Font(color="9C6500")
            elif "abierto" in estado_nombre:
                estado_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rojo claro
                estado_cell.font = Font(color="9C0006")

        # Bordes para todas las celdas
        for col in range(1, 10):
            cell = ws.cell(row=row_num, column=col)
            cell.border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            cell.alignment = Alignment(vertical='center')

        row_num += 1

    # --- AJUSTAR ANCHOS DE COLUMNA ---
    column_widths = [8, 35, 30, 15, 12, 15, 18, 18, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # --- FILTROS AUTOMÁTICOS ---
    ws.auto_filter.ref = f"A5:I{row_num-1}"

    # --- FILA DE TOTALES ---
    total_row = row_num + 1
    ws.cell(row=total_row, column=1, value=f"TOTAL TICKETS: {tickets.count()}")
    ws.cell(row=total_row, column=1).font = Font(bold=True, size=11)
    
    # Merge cells para que se vea completo
    ws.merge_cells(f'A{total_row}:B{total_row}')

    # --- GENERAR RESPUESTA ---
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_tickets.xlsx"'
    wb.save(response)

    return response


@login_required
def reportes_tickets_pdf(request):
    if not require_role(request.user, "ADMIN"):
        return HttpResponseForbidden("No tienes permiso para exportar.")

    tickets = Ticket.objects.select_related(
        "solicitante",
        "estado",
        "prioridad",
        "area_afectada",
    ).order_by("-fecha_creacion")

    context = {
        "tickets": tickets,
        "generado_por": request.user.email,
        "fecha_generacion": timezone.now(),
    }

    pdf_response = render_to_pdf("reportes/tickets_pdf.html", context)

    if pdf_response is None:
        return HttpResponse("Error al generar el PDF.", status=500)

    pdf_response["Content-Disposition"] = 'attachment; filename="reporte_tickets.pdf"'
    return pdf_response

# -------------------------------------------------------------------
# ADMIN: Gestión de Tickets
# -------------------------------------------------------------------

@login_required
def tickets_listar(request):
    if not require_role(request.user, "ADMIN"):
        return HttpResponseForbidden("No tienes permiso para ver este contenido.")

    tickets = (
        Ticket.objects
        .select_related("solicitante", "estado", "prioridad", "area_afectada", "categoria")
        .prefetch_related("asignaciones__tecnico_asignado__usuario")
        .all()
    )

    estado_id = request.GET.get("estado")
    prioridad_id = request.GET.get("prioridad")
    tecnico_id = request.GET.get("tecnico")
    area_id = request.GET.get("area")
    q = request.GET.get("q")

    if estado_id:
        tickets = tickets.filter(estado_id=estado_id)

    if prioridad_id:
        tickets = tickets.filter(prioridad_id=prioridad_id)

    if tecnico_id:
        tickets = tickets.filter(
            asignaciones__tecnico_asignado_id=tecnico_id,
            asignaciones__activo=True,
        )

    if area_id:
        tickets = tickets.filter(area_afectada_id=area_id)

    if q:
        tickets = tickets.filter(
            Q(titulo__icontains=q) | Q(descripcion__icontains=q)
        )

    tickets = tickets.order_by("-fecha_creacion").distinct()

    estados = EstadoTicket.objects.all()
    prioridades = Prioridad.objects.all()
    tecnicos = Tecnico.objects.select_related("usuario").all()
    areas = AreaAfectada.objects.all()

    return render(request, "admin/tickets_listar.html", {
        "tickets": tickets,
        "estados": estados,
        "prioridades": prioridades,
        "tecnicos": tecnicos,
        "areas": areas,
    })


@login_required
def tickets_detalle(request, ticket_id):
    if not require_role(request.user, "ADMIN"):
        return HttpResponseForbidden("No tienes permiso para ver este ticket.")

    ticket = get_object_or_404(
        Ticket.objects.select_related(
            "solicitante", "estado", "categoria", "prioridad", "area_afectada"
        ),
        id=ticket_id,
    )

    categorias = Categoria.objects.all()
    prioridades = Prioridad.objects.all()
    estados = EstadoTicket.objects.all()
    areas = AreaAfectada.objects.all()
    tecnicos = Tecnico.objects.select_related("usuario")

    if request.method == "POST":
        # --- COMENTARIO ---
        if "comentario_texto" in request.POST:
            texto = request.POST.get("comentario_texto", "").strip()
            archivo = request.FILES.get("comentario_archivo")

            if texto:
                ComentarioTicket.objects.create(
                    ticket=ticket,
                    usuario=request.user,
                    texto=texto,
                    archivo=archivo
                )

                # Notificar al solicitante
                Notificacion.objects.create(
                    ticket=ticket,
                    usuario_destino=ticket.solicitante,
                    titulo=f"Nuevo comentario en Ticket #{ticket.id}",
                    mensaje=f"El administrador agregó un comentario."
                )

                messages.success(request, "Comentario agregado correctamente.")
                return redirect("tickets_detalle", ticket_id=ticket.id)

        # --- ACTUALIZACIÓN DE TICKET ---
        nuevo_estado_id = request.POST.get("estado")
        nueva_categoria_id = request.POST.get("categoria")
        nueva_prioridad_id = request.POST.get("prioridad")
        nueva_area_id = request.POST.get("area_afectada")
        nuevo_tecnico_id = request.POST.get("tecnico_asignado")
        comentario = request.POST.get("comentario")

        estado_anterior = ticket.estado

        # --- Actualización de datos básicos ---
        if nueva_categoria_id:
            ticket.categoria_id = int(nueva_categoria_id)

        if nueva_prioridad_id:
            ticket.prioridad_id = int(nueva_prioridad_id)

        if nueva_area_id:
            ticket.area_afectada_id = int(nueva_area_id)

        if nuevo_estado_id:
            ticket.estado_id = int(nuevo_estado_id)

        # --- SLA: si el estado actual es final, marcamos fecha_cierre ---
        if ticket.estado and ticket.estado.es_final:
            if ticket.fecha_cierre is None:
                ticket.fecha_cierre = timezone.now()
        else:
            # Si se reabre el ticket (estado no final), quitamos fecha_cierre
            ticket.fecha_cierre = None

        ticket.save()

        # --- Asignación de técnico ---
        tecnico_asignado = None
        if nuevo_tecnico_id:
            tecnico_asignado = get_object_or_404(Tecnico, id=int(nuevo_tecnico_id))
            AsignacionTicket.objects.update_or_create(
                ticket=ticket,
                defaults={
                    "tecnico_asignado": tecnico_asignado,
                    "activo": True,
                },
            )

        # --- Historial ---
        HistorialTicket.objects.create(
            ticket=ticket,
            usuario=request.user,
            estado_anterior=estado_anterior,
            estado_nuevo=ticket.estado,
            comentario=comentario or "Actualización realizada por el administrador.",
        )

        # --- Notificaciones ---
        if tecnico_asignado:
            Notificacion.objects.create(
                ticket=ticket,
                usuario_destino=tecnico_asignado.usuario,
                tipo_notificacion="asignacion",
                titulo=f"Ticket #{ticket.id} asignado",
                mensaje=f"Se te asignó el ticket: {ticket.titulo}",
                canal_notificacion="portal",
            )

        Notificacion.objects.create(
            ticket=ticket,
            usuario_destino=ticket.solicitante,
            tipo_notificacion="cambio_estado",
            titulo=f"Ticket #{ticket.id} actualizado",
            mensaje=f"El estado actual es: {ticket.estado.nombre_estado}",
            canal_notificacion="portal",
        )

        # --- Mensaje con información de SLA (si aplica) ---
        sla_msg_extra = ""
        if ticket.fecha_cierre and ticket.sla_horas_objetivo:
            duracion_horas = (ticket.fecha_cierre - ticket.fecha_creacion).total_seconds() / 3600
            duracion_horas_redondeado = round(duracion_horas, 1)

            if duracion_horas <= ticket.sla_horas_objetivo:
                sla_msg_extra = f" Cumplió el SLA: {duracion_horas_redondeado}h ≤ {ticket.sla_horas_objetivo}h."
            else:
                sla_msg_extra = f" No cumplió el SLA: {duracion_horas_redondeado}h > {ticket.sla_horas_objetivo}h."

        messages.success(
            request,
            f"El ticket #{ticket.id} se actualizó correctamente.{sla_msg_extra}"
        )
        return redirect("tickets_detalle", ticket_id=ticket.id)

    # --- Cálculo de SLA para la vista (GET) ---
    cumplio_sla = None
    horas_totales = None
    if ticket.fecha_cierre and ticket.sla_horas_objetivo:
        segundos = (ticket.fecha_cierre - ticket.fecha_creacion).total_seconds()
        horas_totales = round(segundos / 3600, 1)
        cumplio_sla = horas_totales <= ticket.sla_horas_objetivo

    # --- Obtener comentarios ---
    comentarios = ticket.comentarios.select_related("usuario").order_by("fecha_creacion")

    return render(request, "admin/tickets_detalle.html", {
        "ticket": ticket,
        "categorias": categorias,
        "prioridades": prioridades,
        "estados": estados,
        "areas": areas,
        "tecnicos": tecnicos,
        "cumplio_sla": cumplio_sla,
        "horas_totales": horas_totales,
        "comentarios": comentarios,
    })


@login_required
def tickets_eliminar(request, ticket_id):
    """Eliminar ticket (solo ADMIN)"""
    if not require_role(request.user, "ADMIN"):
        return HttpResponseForbidden("No tienes permiso para eliminar tickets.")

    ticket = get_object_or_404(Ticket, id=ticket_id)

    if request.method == "POST":
        ticket_numero = ticket.id
        ticket_titulo = ticket.titulo
        ticket.delete()
        messages.success(request, f"Ticket #{ticket_numero} - {ticket_titulo} eliminado correctamente.")
        return redirect("tickets_listar")

    return render(request, "admin/tickets_eliminar.html", {
        "ticket": ticket
    })


# -------------------------------------------------------------------
# TÉCNICO: listado de tickets
# -------------------------------------------------------------------

@login_required
def tickets_tecnico_listar(request):
    if not require_role(request.user, "TECNICO"):
        return HttpResponseForbidden("No tienes permiso para ver este contenido.")

    # Perfil de técnico
    tecnico = getattr(request.user, "tecnico", None)
    if tecnico is None:
        tecnico, _ = Tecnico.objects.get_or_create(usuario=request.user)

    # --- Scope: "mios" (por defecto) o "todos" ---
    scope = request.GET.get("scope", "mios")  # valores: "mios" o "todos"

    # Query base
    base_qs = (
        Ticket.objects
        .select_related("solicitante", "estado", "prioridad", "area_afectada", "categoria")
        .order_by("-fecha_creacion")
    )

    if scope == "todos":
        tickets_qs = base_qs
    else:
        tickets_qs = base_qs.filter(
            asignaciones__tecnico_asignado=tecnico,
            asignaciones__activo=True,
        ).distinct()

    # --- Filtros GET ---
    estado_id = request.GET.get("estado")
    prioridad_id = request.GET.get("prioridad")
    q = request.GET.get("q")

    if estado_id:
        tickets_qs = tickets_qs.filter(estado_id=estado_id)

    if prioridad_id:
        tickets_qs = tickets_qs.filter(prioridad_id=prioridad_id)

    if q:
        tickets_qs = tickets_qs.filter(
            Q(titulo__icontains=q) | Q(descripcion__icontains=q)
        )

    # Evaluamos el queryset y le asignamos la "asignación activa" a cada ticket
    tickets = list(tickets_qs)
    ticket_ids = [t.id for t in tickets]

    asignaciones = AsignacionTicket.objects.filter(
        ticket_id__in=ticket_ids,
        activo=True,
    ).select_related("tecnico_asignado__usuario")

    asignacion_por_ticket = {a.ticket_id: a for a in asignaciones}

    for t in tickets:
        t.asignacion_activa = asignacion_por_ticket.get(t.id)

    # Catálogos
    estados = EstadoTicket.objects.all()
    prioridades = Prioridad.objects.all()

    return render(request, "tecnico/tickets_listar.html", {
        "tickets": tickets,
        "estados": estados,
        "prioridades": prioridades,
        "scope": scope,
        "tecnico": tecnico,
    })
@login_required
def ticket_tecnico_detalle(request, ticket_id):
    if not require_role(request.user, "TECNICO"):
        return HttpResponseForbidden("No tienes permiso para ver este ticket.")

    # Perfil de técnico del usuario logueado
    tecnico = getattr(request.user, "tecnico", None)
    if tecnico is None:
        tecnico, _ = Tecnico.objects.get_or_create(usuario=request.user)

    # Ticket a visualizar
    ticket = get_object_or_404(
        Ticket.objects.select_related(
            "solicitante",
            "estado",
            "prioridad",
            "area_afectada",
            "categoria",
        ),
        id=ticket_id,
    )

    # Asignación ACTIVA (independiente de quién sea el técnico)
    asignacion_global = (
        AsignacionTicket.objects
        .filter(ticket=ticket, activo=True)
        .select_related("tecnico_asignado__usuario")
        .first()
    )

    # ¿El ticket está asignado a ESTE técnico?
    asignado_a_mi = (
        asignacion_global is not None
        and asignacion_global.tecnico_asignado_id == tecnico.id
    )

    puede_editar = asignado_a_mi
    estados = EstadoTicket.objects.all()

    # ---------------- POST (actualizar estado o agregar comentario) ----------------
    if request.method == "POST":
        # --- COMENTARIO ---
        if "comentario_texto" in request.POST:
            texto = request.POST.get("comentario_texto", "").strip()
            archivo = request.FILES.get("comentario_archivo")

            if texto:
                ComentarioTicket.objects.create(
                    ticket=ticket,
                    usuario=request.user,
                    texto=texto,
                    archivo=archivo
                )

                # Notificar al solicitante
                usuario_nombre = request.user.get_full_name() or request.user.email
                Notificacion.objects.create(
                    ticket=ticket,
                    usuario_destino=ticket.solicitante,
                    titulo=f"Nuevo comentario en Ticket #{ticket.id}",
                    mensaje=f"{usuario_nombre} agregó un comentario."
                )

                messages.success(request, "Comentario agregado correctamente.")
                return redirect("ticket_tecnico_detalle", ticket_id=ticket.id)

        # --- CAMBIAR ESTADO ---
        if not puede_editar:
            messages.error(
                request,
                "No puedes modificar un ticket que no está asignado a ti."
            )
            return redirect("ticket_tecnico_detalle", ticket_id=ticket.id)

        nuevo_estado_id = request.POST.get("estado")
        comentario = (request.POST.get("comentario") or "").strip()

        estado_anterior = ticket.estado

        if nuevo_estado_id:
            ticket.estado_id = int(nuevo_estado_id)

        estado_nuevo = ticket.estado
        if estado_nuevo.es_final:
            if ticket.fecha_cierre is None:
                ticket.fecha_cierre = timezone.now()
        else:
            ticket.fecha_cierre = None

        ticket.save()

        HistorialTicket.objects.create(
            ticket=ticket,
            usuario=request.user,
            estado_anterior=estado_anterior,
            estado_nuevo=ticket.estado,
            comentario=comentario or "Actualización realizada por el técnico.",
        )

        messages.success(request, "El ticket se actualizó correctamente.")
        return redirect("ticket_tecnico_detalle", ticket_id=ticket.id)
    
    # ---------------- GET ----------------
    comentarios = ticket.comentarios.select_related("usuario").order_by("fecha_creacion")

    return render(request, "tecnico/ticket_detalle.html", {
        "ticket": ticket,
        "estados": estados,
        "puede_editar": puede_editar,
        "asignacion_global": asignacion_global,
        "asignado_a_mi": asignado_a_mi,
        "comentarios": comentarios,
    })

# -------------------------------------------------------------------
# USUARIO: tickets propios
# -------------------------------------------------------------------

@login_required
def tickets_usuario_listar(request):
    if not require_role(request.user, "USUARIO"):
        return HttpResponseForbidden("No tienes permiso para ver estos tickets.")

    tickets = Ticket.objects.filter(
        solicitante=request.user
    ).select_related("estado", "categoria", "prioridad").order_by("-fecha_creacion")

    return render(request, "usuario/tickets_listar.html", {
        "tickets": tickets
    })


@login_required
def ticket_usuario_crear(request):
    if not require_role(request.user, "USUARIO"):
        return HttpResponseForbidden("No tienes permiso para crear tickets.")

    areas = AreaAfectada.objects.all()

    estado_abierto, _ = EstadoTicket.objects.get_or_create(
        nombre_estado="Abierto",
        defaults={
            "descripcion": "Ticket creado por el usuario",
            "es_final": False,
        },
    )

    if request.method == "POST":
        titulo = request.POST.get("titulo")
        descripcion = request.POST.get("descripcion")
        area_id = request.POST.get("area_afectada")
        archivo = request.FILES.get("archivo")

        if not titulo or not descripcion or not area_id:
            messages.error(request, "Título, descripción y área afectada son obligatorios.")
            return render(request, "usuario/ticket_crear.html", {"areas": areas})

        ticket = Ticket.objects.create(
            titulo=titulo,
            descripcion=descripcion,
            solicitante=request.user,
            area_afectada_id=int(area_id),
            estado=estado_abierto,
            archivo=archivo,
        )

        HistorialTicket.objects.create(
            ticket=ticket,
            usuario=request.user,
            estado_anterior=None,
            estado_nuevo=estado_abierto,
            comentario="Ticket creado por el solicitante.",
        )

        admins = Usuario.objects.filter(rol__nombre_rol__iexact="ADMIN")
        for admin in admins:
            Notificacion.objects.create(
                ticket=ticket,
                usuario_destino=admin,
                tipo_notificacion="creacion",
                titulo=f"Nuevo ticket #{ticket.id}",
                mensaje=f"{request.user.email} creó el ticket: {ticket.titulo}",
                canal_notificacion="portal",
            )

        messages.success(request, f"Tu ticket #{ticket.id} fue creado correctamente.")
        return redirect("tickets_usuario_listar")

    return render(request, "usuario/ticket_crear.html", {"areas": areas})


@login_required
def ticket_usuario_detalle(request, ticket_id):
    if not require_role(request.user, "USUARIO"):
        return HttpResponseForbidden("No tienes permiso para ver este ticket.")

    ticket = get_object_or_404(
        Ticket.objects.select_related("estado", "categoria", "prioridad", "solicitante"),
        id=ticket_id,
        solicitante=request.user,
    )

    # --- Procesar formulario de comentario ---
    if request.method == "POST" and "comentario_texto" in request.POST:
        texto = request.POST.get("comentario_texto", "").strip()
        archivo = request.FILES.get("comentario_archivo")

        if texto:
            comentario = ComentarioTicket.objects.create(
                ticket=ticket,
                usuario=request.user,
                texto=texto,
                archivo=archivo
            )

            # --- Notificar al técnico asignado (si existe) ---
            asignacion = AsignacionTicket.objects.filter(ticket=ticket, activo=True).first()
            if asignacion and asignacion.tecnico_asignado:
                usuario_nombre = request.user.get_full_name() or request.user.email
                Notificacion.objects.create(
                    ticket=ticket,
                    usuario_destino=asignacion.tecnico_asignado.usuario,
                    titulo=f"Nuevo comentario en Ticket #{ticket.id}",
                    mensaje=f"{usuario_nombre} agregó un comentario: {texto[:100]}..."
                )

            messages.success(request, "Comentario agregado correctamente.")
            return redirect("ticket_usuario_detalle", ticket_id=ticket.id)

    # --- Procesar formulario de calificación ---
    if request.method == "POST" and "puntuacion" in request.POST:
        puntuacion = int(request.POST.get("puntuacion"))
        resuelto = request.POST.get("resuelto") == "si"
        comentario_csat = request.POST.get("comentario_csat", "").strip()

        # Crear calificación (solo si no existe)
        CalificacionTicket.objects.get_or_create(
            ticket=ticket,
            defaults={
                "usuario": request.user,
                "puntuacion": puntuacion,
                "resuelto": resuelto,
                "comentario": comentario_csat,
            }
        )

        messages.success(request, "¡Gracias por tu calificación!")
        return redirect("ticket_usuario_detalle", ticket_id=ticket.id)

    # --- Obtener comentarios ---
    comentarios = ticket.comentarios.select_related("usuario").order_by("fecha_creacion")

    # --- Verificar si ya calificó ---
    ya_califico = hasattr(ticket, 'calificacion')

    return render(request, "usuario/ticket_detalle.html", {
        "ticket": ticket,
        "comentarios": comentarios,
        "ya_califico": ya_califico,
    })


# -------------------------------------------------------------------
# EDITAR PERFIL (todos los roles)
# -------------------------------------------------------------------

@login_required
def editar_perfil(request):
    """
    Permite a cualquier usuario (ADMIN, TECNICO, USUARIO) editar:
    - Avatar (ícono de perfil)
    - Contraseña
    """
    if request.method == "POST":
        # --- CAMBIAR AVATAR ---
        if "avatar" in request.POST:
            nuevo_avatar = request.POST.get("avatar")
            if nuevo_avatar in dict(Usuario.AVATAR_CHOICES):
                request.user.avatar = nuevo_avatar
                request.user.save(update_fields=['avatar'])
                messages.success(request, "Avatar actualizado correctamente.")
                return redirect("editar_perfil")

        # --- CAMBIAR CONTRASEÑA ---
        if "password_actual" in request.POST:
            password_actual = request.POST.get("password_actual")
            password_nueva = request.POST.get("password_nueva")
            password_confirmar = request.POST.get("password_confirmar")

            # Validar contraseña actual
            if not request.user.check_password(password_actual):
                messages.error(request, "La contraseña actual es incorrecta.")
                return redirect("editar_perfil")

            # Validar que coincidan
            if password_nueva != password_confirmar:
                messages.error(request, "Las contraseñas nuevas no coinciden.")
                return redirect("editar_perfil")

            # Validar longitud mínima
            if len(password_nueva) < 8:
                messages.error(request, "La contraseña debe tener al menos 8 caracteres.")
                return redirect("editar_perfil")

            # Actualizar contraseña
            request.user.set_password(password_nueva)
            request.user.save()
            messages.success(request, "Contraseña actualizada correctamente. Por favor, inicia sesión nuevamente.")
            return redirect("login")

    return render(request, "editar_perfil.html", {
        "avatares": Usuario.AVATAR_CHOICES,
    })


# -------------------------------------------------------------------
# NOTIFICACIONES (panel)
# -------------------------------------------------------------------

@login_required
def notificaciones_listar(request):
    notificaciones = Notificacion.objects.filter(
        usuario_destino=request.user
    ).order_by("-fecha_envio")

    return render(request, "notificaciones/listar.html", {
        "notificaciones": notificaciones
    })


@login_required
def notificacion_marcar_leida(request, notif_id):
    notif = get_object_or_404(Notificacion, id=notif_id, usuario_destino=request.user)
    notif.leida = True
    notif.save()
    return redirect("notificaciones_listar")
